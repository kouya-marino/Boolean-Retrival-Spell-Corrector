'''
@Prashant Rawat 
Titile: Boolean Operator with spell check and wild card search
'''
import re
import nltk
from collections import defaultdict
from openpyxl import load_workbook


dic = defaultdict(list)
wb2 = load_workbook('English.xlsx')
source = wb2.active
d = wb2['Sheet1']['A']
d2 = defaultdict(list)
set2=set()
stopwords = nltk.corpus.stopwords.words('english')
corpus_words=[]
for i in range(len(d)) :
    c = source.cell(row = i + 1,column = 1).value
    text = re.sub("[^a-zA-Z]+", " ", c)
    tokens = nltk.tokenize.word_tokenize(text)
    #print(tokens)
    stemmer = nltk.stem.porter.PorterStemmer()
    stopwords = nltk.corpus.stopwords.words('english')
    s = []
    #xxx= []
    for token in tokens :
        #print(token)
        token = token.lower()
        #xxx.append(token)
        if token not in stopwords and len(token) > 2 :
            corpus_words.append(token)
            s .append(stemmer.stem(token))
            
    
    s = set(s)
    #xxx=set(xxx)
    for k in s:
        dic[k].append(i+1)
'''
###### SAVE Inverted Index ########

if you want to save Index uncomment following 4 lines
'''    
    #fo = open('Inverted_Index.txt', "w")
    #for k, v in dic.items():
     #   fo.write(str(k) + '-> '+ str(v) + '\n\n')
    #fo.close()   

corpus_words=set(corpus_words)    

'''def Not(w1,w2):
    #print("Not")
    All=set(range(1,2273))
    ans =All.difference(w2.union(w1))
        
    if len(ans) > 0 :
        print(ans)
    
    else:
        print("NO Match Found")
    
    #print(ans)

def Or(w1,w2):
    #print("Or")
    ans = w2.union(w1)
        
    if len(ans) > 0 :
        print(ans)
    
    else:
        print("NO Match Found")
    
    #print(ans)
'''
def And(w1,w2):    
    #print("And")
          
    ans = w2.intersection(w1)
        
    if len(ans) > 0 :
        print(ans)
    
    else:
        print("NO Match Found")
        
        
        



#i = input()
f = open('query_eng.txt','r')
for i in f:
    
    text = re.sub("[^a-zA-Z]+", " ", i)
    tokens = nltk.tokenize.word_tokenize(text)
    stemmer = nltk.stem.porter.PorterStemmer()

    ss = [] ;
    op = "and"

    '''

    ################### Normal implementation with stemming only##################
    
'''

    for token in tokens:
        token.lower()
        if token not in stopwords:
            ss.append(stemmer.stem(token))

    word1=ss[0]
    word2=ss[1]
    w1 = set(dic[word1])
    w2 = set(dic[word2])

    '''
    ### Boolean operation 
    '''

    print("With only Stemming\n")
    And(w1,w2)
    print("\n\n")



    '''
    ############## With Spell Correction #########################################

    '''

    ###############'''
    print("With Spell Correction\n")
    '''
    #
    '''
    def lcsf(a,b):
        la=len(a)
        lb=len(b)

        l=[[0 for i in range(0,la+1)]for i in range(0,lb+1)]
        for i in l:
            #print i
            for i in range(1,lb+1):
                for j in range(1,la+1):
                    if(a[j-1]==b[i-1]):
     
                        l[i][j]=l[i-1][j-1]+1
                    else:
                        if l[i][j-1]>=l[i-1][j]:
                            l[i][j]=l[i][j-1]
                        else: #l[i][j-1]>l[i-1][j]:
                            l[i][j]=l[i-1][j]
                #else:
                 #   print "sths wrong"
                    
        return l[lb][la]    
    
    wt1=0.0;
    n_word1=""
    n_word2=""
    for i in corpus_words:
        lcs=lcsf(i,word1)
        m=len(i)
        n=len(word1)
        n_wt=float(2*lcs)/float((m+n))
        if n_wt > wt1:
            wt1=n_wt
            n_word1=i

    sc_word1=stemmer.stem(n_word1)        
    wt2=0.0        
    for j in corpus_words:
        lcs=lcsf(j,word2)
        m=len(j)
        n=len(word2)
        n_wt=float(2*lcs)/float((m+n))
        if n_wt > wt2:
            wt2=n_wt
            n_word2=j
        
    print(n_word1,n_word2)
    sc_word2=stemmer.stem(n_word2)

    '''
    ###### SAVE QUERY.txt ###########
    '''
    fl1=str(word1)+".txt"
    fl2=str(word2)+".txt"
    sc_w1 = set(dic[sc_word1])
    sc_w2 = set(dic[sc_word2])

    if wt1 < 1.00 and wt2 == 1.00:
        try:
            p=open('query/'+str(word1)+" "+str(word2)+".txt",'r')
            
        except:
            with open('query/'+fl1, "a") as myfile:
                myfile.write(word1)
                myfile.write(" ")
                myfile.write(word2)
                myfile.write("->  ")
                myfile.write(n_word1)
                myfile.write(" ")
                myfile.write(n_word2)
                myfile.write("\n\n")
                myfile.close()
            with open('OUT_ENGLISH.txt', "a") as myf:
                myf.write(word1)
                myf.write(" ")
                myf.write(word2)
                myf.write("->  ")
                myf.write(n_word1)
                myf.write(" ")
                myf.write(n_word2)
                myf.write("\n\n")
                myf.close()
        p.close()
    elif wt2 < 1.00:
        #xxx=str(word1)+" "+str(word2)+".txt"
        try:
            q=open('query/'+str(word1)+" "+str(word2)+".txt",'r')
        except:
            with open('query/'+str(word1)+" "+str(word2)+".txt", "a") as myfile:
                myfile.write(word1)
                myfile.write(" ")
                myfile.write(word2)
                myfile.write("->")
                myfile.write(n_word1)
                myfile.write(" ")
                myfile.write(n_word2)
                myfile.write("\n\n")
                myfile.close()
            with open('OUT_ENGLISH.txt', "a") as myf:
                myf.write(word1)
                myf.write(" ")
                myf.write(word2)
                myf.write("->  ")
                myf.write(n_word1)
                myf.write(" ")
                myf.write(n_word2)
                myf.write("->  ")
                myf.write(str(sc_w2.intersection(sc_w1)))
                myf.write("\n\n")
                myf.close()
    #q.close()
    
   
    sc_w1 = set(dic[sc_word1])
    sc_w2 = set(dic[sc_word2])

    '''

    # Boolean operation

    '''

    And(sc_w1,sc_w2)
    
    print("\n\n")






